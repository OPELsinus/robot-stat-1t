# # from openpyxl import Workbook, load_workbook
# #
# # # Create a new workbook and select the active sheet
# # workbook = load_workbook(r'C:\Users\Abdykarim.D\Downloads\Петропавловск_пример_для_аналитики.xlsx')
#
#
# #
# # sheet_names = workbook.sheetnames
# #
# # # Print the sheet names
# # for sheet_name in sheet_names:
# #
# #     sheet = workbook[sheet_name]
# #
# #     for r in [(19, 28), (28, 36), (36, 51), (51, 58), (58, 67), (67, 86), (86, 98), (98, 123), (123, 126), (126, 136),
# #     (136, 145), (145, 150), (150, 157), (157, 160), (160, 167), (167, 172), (172, 177), (177, 181), (181, 192),
# #     (192, 198), (198, 205), (210, 219), (219, 229), (229, 239), (239, 243), (243, 250), (250, 257), (257, 262), (262, 274), (274, 277), (277, 284), (284, 291), (291, 295), (295, 301), (301, 304),\
# #     (304, 307), (307, 309), (309, 311), (311, 320), (320, 324), (324, 330), (330, 337),
# #     (342, 351), (351, 361), (361, 375), (375, 383), (383, 390), (390, 407), (407, 420), (420, 444),
# #     (444, 447), (447, 459), (459, 467), (467, 471), (471, 479), (479, 485), (485, 489), (489, 494),
# #     (494, 501), (501, 505), (505, 521), (521, 537), (537, 542), (542, 549), (550, 559), (562, 565), (568, 571),
# #               (571, 573), (573, 579)]:
# #         for row in range(r[0] + 1, r[1]):
# #             sheet.row_dimensions[row].outline_level = 1
# #
# # workbook.save("grouped_rows.xlsx")
# #
# #
# #
# #
# #

import os
from copy import copy

from openpyxl import load_workbook
from openpyxl.formula import translate

import xlwings as xw

# upper_part_1 = {
#     'formula1': '=RC[-1]',
#     'formula2': '=(RC[-1]+RC[-3])/2',
#     'formula3': '=(RC[-1]+RC[-3]+RC[-5])/3',
#     'formula4': '=(RC[-1]+RC[-3]+RC[-5]+RC[-7])/4',
# }
#
# upper_part_2 = {
#     'formula1': '=RC[-1]',
#     'formula2': '=RC[-2]+RC[-1]',
#     'formula3': '=RC[-2]+RC[-1]',
#     'formula4': '=RC[-2]+RC[-1]',
# }
#
# down_part_1 = {
#     'formula1': '=RC[-1]',
#     'formula2': '=R[1]C[-2]',
#     'formula3': '=R[1]C[-2]',
#     'formula4': '=R[1]C[-2]',
# }
#
# down_part_2 = {
#     'formula1': '=RC[-1]',
#     'formula2': '=RC[-2]+RC[-1]',
#     'formula3': '=RC[-2]+RC[-1]',
#     'formula4': '=RC[-2]+RC[-1]',
# }
#
# down_part_3 = {
#     'formula1': '=RC[-1]',
#     'formula2': '=RC[-2]+RC[-1]',
#     'formula3': '=RC[-2]+RC[-1]',
#     'formula4': '=RC[-2]+RC[-1]',
# }
#
#
# for file in os.listdir(r'\\172.16.8.87\d\.rpa\.agent\robot-stat-1t\Файлы сбора'):
#
#     os.system('taskkill /im excel.exe /f')
#
#     book = load_workbook(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-stat-1t\Файлы сбора', file))
#
#     for sheet_ in book.sheetnames:
#
#         sheet = book[sheet_]
#
#         for i in range(10):
#
#             if 'наименование показателей' not in str(sheet['A1'].value).lower():
#
#                 sheet.delete_rows(1)
#
#     for sheet_ in book.sheetnames:
#
#         sheet = book[sheet_]
#
#         print(f'{file} | {sheet} | {sheet["A1"].value}')
#
#     book.save(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-stat-1t', file))
#     book.close()
#
#     excel_app = xw.App(visible=False)
#     workbook = excel_app.books.open(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-stat-1t', file), corrupt_load=True)
#
#     # wb = xw.Book(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-stat-1t', file), corrupt_load=True)
#
#     print('-----')
#
#     for sheet_ in workbook.sheets:
#
#         print('SHEET', sheet_)
#
#         app = xw.apps.active.books.active.sheets(sheet_)
#
#         for ind, col in enumerate('CEGI'):
#             for row_ in [[3, 9]]:
#                 for row in range(row_[0], row_[1]):
#                     # sheet[f'{col}{row}'].value = None
#                     try:
#                         app.range(f'{col}{row}').formula = upper_part_1.get(f'formula{ind + 1}')
#                     except Exception as err:
#                         print(f'BROKE | {col}{row} | {upper_part_1.get(f"formula{ind + 1}")} | {sheet_} --- {err}')
#                         break
#                     # sheet[f'{col}{row}'].data_type = 'formula'
#                     print(f'{col}{row} | {upper_part_1.get(f"formula{ind + 1}")} | {sheet_}')
#
#     workbook.save(os.path.join(r'\\172.16.8.87\d\.rpa\.agent\robot-stat-1t', file))
#
#     os.system('taskkill /im excel.exe /f')
#
# #


import openpyxl
import os


def copy_sheet_contents(source_sheet, target_sheet):
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet.cell(
                row=cell.row, column=cell.column, value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)


def replace_sheets_with_template(template_path, target_files):
    # Load the template workbook and sheet
    template_wb = openpyxl.load_workbook(template_path)
    template_sheet = template_wb.active

    for file_path in target_files:
        target_wb = openpyxl.load_workbook(file_path)

        sheet_names = target_wb.sheetnames
        for sheet_name in sheet_names:
            # Remove the existing sheet
            target_wb.remove(target_wb[sheet_name])
            # Create a new sheet with the same name
            new_sheet = target_wb.create_sheet(title=sheet_name)
            # Copy contents from template to the new sheet
            copy_sheet_contents(template_sheet, new_sheet)

        # Save the modified workbook
        target_wb.save(file_path)
        print(f"Processed {file_path}")


target_files = []

for file in os.listdir(r"\\172.16.8.87\d\.rpa\.agent\robot-stat-1t\Файлы сбора"):
    target_files.append(
        os.path.join(r"\\172.16.8.87\d\.rpa\.agent\robot-stat-1t\Файлы сбора", file)
    )


template_path = r"\\172.16.8.87\d\.rpa\.agent\robot-stat-1t\1Т шаблон.xlsx"

replace_sheets_with_template(template_path, target_files)

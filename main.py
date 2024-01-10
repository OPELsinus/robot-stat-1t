import datetime
import os
import random
import re
import shutil
import time
from contextlib import suppress
from math import floor
from time import sleep
import pandas as pd

import psycopg2
import xlrd
from mouseinfo import screenshot
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from pywinauto import keyboard
import xlwings as xw

from config import logger, tg_token, chat_id, db_host, robot_name, db_port, db_name, db_user, db_pass, ip_address, saving_path, saving_path_1c, download_path, ecp_paths, main_excel_files, adb_db_password, adb_db_name, adb_db_username, adb_ip, adb_port, mapping_file, filled_files, reports_saving_path
from core import Odines
from tools.app import App
from tools.web import Web


def sql_create_table():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
        CREATE TABLE IF NOT EXISTS ROBOT.{robot_name.replace("-", "_")} (
            started_time timestamp,
            ended_time timestamp,
            store_name text UNIQUE,
            short_name text,
            executor_name text,
            status text,
            status_1c text,
            error_reason text,
            error_saved_path text,
            execution_time text,
            ecp_path text
            )
        '''
    c = conn.cursor()
    c.execute(table_create_query)

    conn.commit()
    c.close()
    conn.close()


def delete_by_id(id):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
                DELETE FROM ROBOT.{robot_name.replace("-", "_")} WHERE id = '{id}'
                '''
    c = conn.cursor()
    c.execute(table_create_query)
    conn.commit()
    c.close()
    conn.close()


def get_all_data():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            order by started_time asc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['started_time', 'ended_time', 'full_name', 'short_name', 'executor_name', 'status', 'status_1c', 'error_reason', 'error_saved_path', 'execution_time', 'ecp_path']

    cur.close()
    conn.close()

    return df1


def get_data_by_name(store_name):
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            where store_name = '{store_name}'
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())
    # df1.columns = ['started_time', 'ended_time', 'store_id', 'name', 'status', 'error_reason', 'error_saved_path', 'execution_time']

    cur.close()
    conn.close()

    return len(df1)


def get_data_to_execute():
    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)
    table_create_query = f'''
            SELECT * FROM ROBOT.{robot_name.replace("-", "_")}
            where (status_1c != 'success' and status_1c != 'processing')
            and (executor_name is NULL or executor_name = '{ip_address}')
            order by started_time desc
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df1 = pd.DataFrame(cur.fetchall())

    with suppress(Exception):
        df1.columns = ['started_time', 'ended_time', 'full_name', 'short_name', 'executor_name', 'status', 'status_1c', 'error_reason', 'error_saved_path', 'execution_time', 'ecp_path']

    cur.close()
    conn.close()

    return df1


def insert_data_in_db(started_time: str, store_name: str, short_name: str, executor_name: str, status_: str, status_1c: str, error_reason: str, error_saved_path: str, execution_time: int, ecp_path_: str):

    conn = psycopg2.connect(host=db_host, port=db_port, database=db_name, user=db_user, password=db_pass)

    print('Started inserting')
    # query_delete_id = f"""
    #         delete from ROBOT.{robot_name.replace("-", "_")}_2 where store_id = '{store_id}'
    #     """
    query_delete = f"""
        delete from ROBOT.{robot_name.replace("-", "_")} where store_name = '{store_name}'
    """
    query = f"""
        INSERT INTO ROBOT.{robot_name.replace("-", "_")} (started_time, ended_time, store_name, short_name, executor_name, status, status_1c, error_reason, error_saved_path, execution_time, ecp_path)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """
    # ended_time = '' if status_ != 'success' else datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    ended_time = datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f")
    values = (
        started_time,
        ended_time,
        store_name,
        short_name,
        executor_name,
        status_,
        status_1c,
        error_reason,
        error_saved_path,
        str(execution_time),
        ecp_path_
    )

    # print(values)

    cursor = conn.cursor()

    cursor.execute(query_delete)
    # conn.autocommit = True
    try:
        cursor.execute(query_delete)
        # cursor.execute(query_delete_id)
    except Exception as e:
        print('GOVNO', e)
        pass
    if True:
        cursor.execute(query, values)
    # except Exception as e:
    #     conn.rollback()
    #     print(f"Error: {e}")

    conn.commit()

    cursor.close()
    conn.close()


def get_all_branches_with_codes():

    conn = psycopg2.connect(dbname=adb_db_name, host=adb_ip, port=adb_port,
                            user=adb_db_username, password=adb_db_password)

    cur = conn.cursor(name='1583_first_part')

    query = f"""
        select db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name 
        from dwh_data.dim_branches db
        left join dwh_data.dim_store ds on db.id_sale_object = ds.sale_source_obj_id
        where ds.store_name like '%Торговый%' and current_date between ds.datestart and ds.dateend
        group by db.id_sale_object, ds.source_store_id, ds.store_name, ds.sale_obj_name
        order by ds.source_store_id
    """

    cur.execute(query)

    print('Executed')

    df1 = pd.DataFrame(cur.fetchall())
    df1.columns = ['branch_id', 'store_id', 'store_name', 'store_normal_name']

    cur.close()
    conn.close()

    return df1


def sign_ecp(ecp):
    logger.info('Started ECP')
    logger.info(f'KEY: {ecp}')
    app = App('')

    el = {"title": "Открыть файл", "class_name": "SunAwtDialog", "control_type": "Window",
          "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}

    if app.wait_element(el, timeout=30):

        keyboard.send_keys(ecp.replace('(', '{(}').replace(')', '{)}'), pause=0.01, with_spaces=True)
        sleep(0.05)
        keyboard.send_keys('{ENTER}')

        if app.wait_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                             "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}, timeout=30):
            app.find_element({"title_re": "Формирование ЭЦП.*", "class_name": "SunAwtDialog", "control_type": "Window",
                              "visible_only": True, "enabled_only": True, "found_index": 0, "parent": None}).type_keys('Aa123456')

            sleep(2)
            keyboard.send_keys('{ENTER}')
            sleep(3)
            keyboard.send_keys('{ENTER}')
            app = None
            logger.info('Finished ECP')

        else:
            logger.info('Quit mazafaka1')
            app = None
            return 'broke'
    else:
        logger.info('Quit mazafaka')
        app = None
        return 'broke'


def save_screenshot(store):
    scr = screenshot()
    save_path = os.path.join(saving_path, 'Ошибки 1Т')
    scr_path = str(os.path.join(os.path.join(saving_path, 'Ошибки 1Т'), str(store + '.png')))
    scr.save(scr_path)

    return scr_path


def wait_loading(web, xpath):
    print('Started loading')
    ind = 0
    element = ''
    while True:
        try:
            print(web.get_element_display('//*[@id="loadmask-1315"]'))
            if web.get_element_display('//*[@id="loadmask-1315"]') == '':
                element = ''
            if (element == '' and web.get_element_display('//*[@id="loadmask-1315"]') == 'none') or (ind >= 500):
                print('Loaded')
                sleep(0.5)
                break
        except:
            print('No loader')
            break
        ind += 1
        sleep(0.05)


def send_file_to_tg(tg_token, chat_id, param, param1):
    pass


def create_and_send_final_report():
    df = get_all_data()

    df.columns = ['Время начала', 'Время окончания', 'Название филиала', 'Короткое название', 'Статус', 'Причина ошибки', 'Пусть сохранения скриншота', 'Время исполнения (сек)', 'Факт1', 'Факт2', 'Факт3', 'Сайт1', 'Сайт2', 'Сайт3']

    df['Время исполнения (сек)'] = df['Время исполнения (сек)'].astype(float)
    df['Время исполнения (сек)'] = df['Время исполнения (сек)'].round()

    df.to_excel('result.xlsx', index=False)

    workbook = load_workbook('result.xlsx')
    sheet = workbook.active

    red_fill = PatternFill(start_color="FFA864", end_color="FFA864", fill_type="solid")
    green_fill = PatternFill(start_color="A6FF64", end_color="A6FF64", fill_type="solid")

    for cell in sheet['D']:
        if cell.value == 'failed':
            cell.fill = red_fill
        if cell.value == 'success':
            cell.fill = green_fill

    for col in 'ABCDGH':

        max_length = max(len(str(cell.value)) for cell in sheet[col])

        if col == 'A' or col == 'B':
            max_length -= 3
        if col == 'D':
            max_length += 5
        if col == 'A':
            max_length -= 3

        sheet.column_dimensions[col].width = max_length

    for col in 'ABCDGEFGH':
        for cell in sheet[col]:
            cell.alignment = Alignment(horizontal='center')

    workbook.save('result.xlsx')

    send_file_to_tg(tg_token, chat_id, 'Отправляем отчёт по заполнению', 'result.xlsx')


def wait_image_loaded(name):

    found = False
    for i in range(100):
        for file in os.listdir(download_path):
            if '.jpg' in file and 'crdownload' not in file:
                shutil.move(os.path.join(download_path, file), os.path.join(reports_saving_path, name + '.jpg'))
                print(file)
                found = True
                break
        sleep(5)
        if found:
            break
    if not found:
        logger.warning('Image not loaded')
        raise Exception('Image not loaded')


def save_and_send(web, save, ecp_sign):

    print('Saving and Sending')
    sleep(1000)
    if save:
        web.execute_script_click_xpath("//span[text() = 'Сохранить']")
        sleep(1)
        print('Clicked Save')
        if web.wait_element("//span[text() = 'Сохранить отчет и Удалить другие']", timeout=5):
            web.execute_script_click_xpath("//span[text() = 'Сохранить отчет и Удалить другие']")

    print('Clicking Send')
    errors = web.find_elements('//*[@id="statflc"]//a', timeout=15)
    errors_count = 0
    for errorik in errors:
        print(f"ERRORS IN STAT: {errorik.get_attr('title').lower()}")
        if 'допустимый' not in errorik.get_attr('title').lower():
            errors_count += 1

    if errors_count == 0:

        print('ALL GOOD')
        web.execute_script_click_xpath("//span[text() = 'Отправить']")
        print('Clicked Send')
        if web.wait_element("//input[@value = 'Персональный компьютер']", timeout=60):
            web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")
        else:
            web.find_element("//button[@class='btn-savesigned ui-button ui-widget ui-state-default ui-corner-all ui-button-text-icon-primary']/span[text() = 'Отправить']").click()
            web.wait_element("//input[@value = 'Персональный компьютер']", timeout=120)
            web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")
        print('Checkpoint on signing')
        sign_ecp(ecp_sign)

        if web.wait_element("//span[text() = 'Продолжить']", timeout=10):
            web.execute_script_click_xpath("//span[text() = 'Продолжить']")

        if web.wait_element("//h1[contains(text(), 'Whitelabel')]", timeout=5):
            for _ in range(10):
                try:
                    web.execute_script_click_xpath("//span[text() = 'Сохранить']")
                except:
                    sleep(60)
                    web.execute_script_click_xpath("//span[text() = 'Сохранить']")
                sleep(1)
                if web.wait_element("//span[text() = 'Сохранить отчет и Удалить другие']", timeout=30):
                    web.execute_script_click_xpath("//span[text() = 'Сохранить отчет и Удалить другие']")
                web.execute_script_click_xpath("//button[@class='btn-savesigned ui-button ui-widget ui-state-default ui-corner-all ui-button-text-icon-primary']/span[text() = 'Отправить']")

                if web.wait_element("//input[@value = 'Персональный компьютер']", timeout=60):
                    web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")
                else:
                    web.find_element("//button[@class='btn-savesigned ui-button ui-widget ui-state-default ui-corner-all ui-button-text-icon-primary']/span[text() = 'Отправить']").click()
                    web.wait_element("//input[@value = 'Персональный компьютер']", timeout=120)
                    web.execute_script_click_xpath("//input[@value = 'Персональный компьютер']")
                print('Checkpoint on signing')
                sign_ecp(ecp_sign)

                if web.wait_element("//span[text() = 'Продолжить']", timeout=10):
                    web.execute_script_click_xpath("//span[text() = 'Продолжить']")

                if not web.wait_element("//h1[contains(text(), 'Whitelabel')]", timeout=5):
                    break
    else:
        print('GOVNO OSHIBKA VYLEZLA')
        sleep(10000)
        raise Exception('ERROR IN EXCEL: Разные цифры')


def wait_loading_1t(web, store):

    for i in range(5):

        if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
            web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")

        if web.wait_element("(//tr[@role='row'])[1]", timeout=5):

            if web.wait_element("//div[contains(text(), '1-Т (кварт')]", timeout=3):
                web.find_element("//div[contains(text(), '1-Т (кварт')]").click()

                return 0

            else:
                saved_path = save_screenshot(store)
                web.close()
                web.quit()

                print('Return those shit')
                return ['Нет 1-Т', saved_path]

        else:
            web.driver.refresh()

    saved_path = save_screenshot(store)
    web.close()
    web.quit()

    print('Return those shit')
    return ['Нет 1-инвест', saved_path]


def proverka_ecp(web):

    if web.wait_element('//*[@id="AgreeId_header_hd-textEl"]', timeout=.5):
        web.execute_script_click_xpath("//span[text() = 'Согласен']")


def start_single_branch(branch_name: str,  values_first_part: dict, values_second_part: dict):

    def pass_later():
        if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
            print('PASSING LATER')
            web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")

    print('Started web')

    ecp_auth = ''
    ecp_sign = ''
    for file in os.listdir(os.path.join(ecp_paths, branch)):

        if 'AUTH' in file:
            ecp_auth = os.path.join(os.path.join(ecp_paths, branch), file)
        if 'GOST' in file:
            ecp_sign = os.path.join(os.path.join(ecp_paths, branch), file)
    print(ecp_auth, '|', ecp_sign)
    web = Web()
    web.run()
    web.get('https://cabinet.stat.gov.kz/')
    logger.info('Check-1')

    logger.info('refreshed')

    proverka_ecp(web=web)

    web.wait_element('//*[@id="idLogin"]')
    web.find_element('//*[@id="idLogin"]').click()

    proverka_ecp(web=web)

    # * --- deprecated (maybe useful in future)
    # web.wait_element('//*[@id="button-1077-btnEl"]')
    # web.find_element('//*[@id="button-1077-btnEl"]').click()
    # * ---
    # proverka_ecp(web=web)
    print()
    # web.wait_element('//*[@id="lawAlertCheck"]')
    # web.find_element('//*[@id="lawAlertCheck"]').click()
    web.execute_script_click_xpath("//input[@id='lawAlertCheck']")

    time.sleep(0.5)
    web.find_element('//*[@id="loginButton"]').click()

    logger.info('Check-2')

    time.sleep(1)

    # send_message_to_tg(tg_token, chat_id, f"Started ECP, {datetime.datetime.now()}")
    sign_ecp(ecp_auth)
    # send_message_to_tg(tg_token, chat_id, f"Finished ECP, {datetime.datetime.now()}")

    logged_in = web.wait_element('//*[@id="idLogout"]/a')

    store = branch.split('\\')[-1]
    # sleep(1000)
    if logged_in:
        if web.find_element("//a[text() = 'Выйти']"):

            if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5):
                try:
                    web.find_element("//span[contains(text(), 'Пройти позже')]").click()
                except:
                    save_screenshot(store)
                    # print('HUETA')
                    # sleep(200)
            logger.info('Check0')
            if web.wait_element('//*[@id="dontAgreeId-inputEl"]', timeout=5):
                web.find_element('//*[@id="dontAgreeId-inputEl"]').click()
                sleep(0.3)
                web.find_element('//*[@id="saveId-btnIconEl"]').click()
                sleep(1)

                # * --- Deprecated (maybe useful)
                # web.find_element('//*[@id="ext-gen1893"]').click()
                # web.find_element('//*[@id="boundlist-1327-listEl"]/ul/li').click()
                # * ---

                web.wait_element('//*[@id="keyCombo-inputEl"]')

                web.execute_script_click_xpath("//*[@id='keyCombo-inputEl']/../following-sibling::td//div")

                web.find_element("//li[contains(text(), 'Персональный компьютер')]").click()
                sleep(1.5)

                web.execute_script_click_xpath("//span[contains(text(), 'Продолжить')]")

                print('Done lol')
                sign_ecp(ecp_sign)
                print('Finished done lol')
                try:
                    if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=5):
                        web.find_element("//span[contains(text(), 'Пройти позже')]").click()

                except:
                    pass

            # web.wait_element('//*[@id="radio-1131-boxLabelEl"]')

            pass_later()
            print('OTCHETY')
            web.wait_element("//span[contains(text(), 'Мои отчёты')]")
            web.execute_script_click_xpath("//span[contains(text(), 'Мои отчёты')]")

            # ? Check if 1Т exists

            pass_later()

            # * ------- Uncomment -------
            wait_loading_1t(web, store)
            # for _ in range(5):
            #
            #     is_loaded = True if len(web.find_elements("//div[contains(@class, 'x-grid-row-expander')]", timeout=15)) >= 1 else False
            #
            #     if is_loaded:
            #         if web.wait_element("//div[contains(text(), '1-Т')]", timeout=3):
            #             web.find_element("//div[contains(text(), '1-Т')]").click()
            #
            #         else:
            #             saved_path = save_screenshot(branch_name)
            #             web.close()
            #             web.quit()
            #
            #             print('Return those shit')
            #             return ['failed', saved_path, 'Нет 1-Т']
            #
            #     else:
            #         web.refresh()
            #
            # if web.wait_element("//span[contains(text(), 'Пройти позже')]", timeout=1.5):
            #     web.execute_script_click_xpath("//span[contains(text(), 'Пройти позже')]")

            sleep(0.5)

            web.find_element('//*[@id="createReportId-btnIconEl"]').click()

            sleep(1)

            # ? Switch to the second window
            print('kek1')
            sleep(7)
            print('switched')
            web.driver.switch_to.window(web.driver.window_handles[-1])

            web.find_element('/html/body/div[1]').click()
            web.wait_element('//*[@id="td_select_period_level_1"]/span')
            web.execute_script_click_js("#btn-opendata")
            sleep(0.3)

            if web.get_element_display('/html/body/div[7]') == 'block':

                web.find_element('/html/body/div[7]/div[11]/div/button[2]').click()

                saved_path = save_screenshot(branch_name)
                web.close()
                web.quit()

                print('Return that shit')
                return ['failed', saved_path, 'Выскочила ошиПочка']

            web.wait_element('//*[@id="sel_statcode_accord"]/div/p/b[1]', timeout=100)
            web.execute_script_click_js("body > div:nth-child(16) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1) > span")

            web.wait_element('//*[@id="sel_rep_accord"]/h3[1]/a')

            # ? Open new report to fill it

            print('Clicking1')
            # web.execute_script_click_js("body > div:nth-child(18) > div.ui-dialog-buttonpane.ui-widget-content.ui-helper-clearfix > div > button:nth-child(1)")
            web.execute_script_click_xpath('/html/body/div[17]/div[11]/div/button[1]/span')

            # ? First page
            print('kek')
            web.wait_element("//a[contains(text(), 'Страница 1')]", timeout=10)
            web.find_element("//a[contains(text(), 'Страница 1')]").click()
            print()

            for row, values in values_first_part.items():

                if values[0] is not None and (float(values[0]) > 0):
                    web.find_element(f'(//*[@id="{row}"]/td[3])[1]').click()
                    web.find_element(f'(//*[@id="{row}_col_2"])').type_keys(round(values[0], 1))

                if values[1] is not None and (float(values[1]) > 0):
                    web.find_element(f'(//*[@id="{row}"]/td[4])[1]').click()
                    web.find_element(f'(//*[@id="{row}_col_3"])').type_keys(round(values[1], 1))

            keyboard.send_keys('{TAB}')
            # sleep(100)
            # ? Second page
            web.wait_element("//a[contains(text(), 'Страница 2')]", timeout=10)
            web.find_element("//a[contains(text(), 'Страница 2')]").click()

            web.find_element('//*[@id="rtime"]').select('1')
            sleep(1)
            print('-----')

            for row, values in values_second_part.items():

                if values[0] is not None and (float(values[0]) > 0):
                    web.find_element(f'(//*[@id="{row}"]/td[3])[2]').click()
                    print(f'Writing {round(values[0], 1)} at {row} | {values}')
                    web.find_element(f'(//*[@id="{row}_col_2"])').type_keys(round(values[0], 1))
                    sleep(10)

                if values[1] is not None and (float(values[1]) > 0):
                    web.find_element(f'(//*[@id="{row}"]/td[4])[2]').click()
                    print(f'Writing {round(values[1], 1)} at {row} | {values}')
                    web.find_element(f'(//*[@id="{row}_col_3"])').type_keys(round(values[1], 1))
                    sleep(10)

            keyboard.send_keys('{TAB}')
            # ? Last page
            web.find_element("//a[contains(text(), 'Данные исполнителя')]").click()
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_3_0']", value='Естаева Акбота Канатовна')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_3_1']", value='7273391350')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_3_2']", value='7073882688')
            web.execute_script(element_type="value", xpath="//*[@id='inpelem_3_3']", value='Yestayeva@magnum.kz')

            # sleep(3000)
            save_and_send(web, save=True, ecp_sign=ecp_sign)
            # sleep(3000)
            # sign_ecp(ecp_sign)
            # sleep(1000)

            wait_image_loaded(branch_name)

            web.close()
            web.quit()

            print('Successed')
            return ['success', '', '']

            # return ['success', '', sites]

    else:

        saved_path = save_screenshot(branch_name)

        web.close()
        web.quit()

        print('Srok istek')
        return ['failed', saved_path, 'Срок ЭЦП истёк']


def get_all_branches():
    conn = psycopg2.connect(host=adb_ip, port=adb_port, database=adb_db_name, user=adb_db_username, password=adb_db_password)
    table_create_query = f'''
                select name_1c_zup from dwh_data.dim_store
                where current_date between datestart and dateend
                group by name_1c_zup
                '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    return pd.DataFrame(cur.fetchall())


def replacements(ind, line):

    if ind == 0:
        return line.replace('С', 'C')
    if ind == 1:
        return line.replace('ТОО ', 'ТОО')
    if ind == 2:
        return line.replace('г. ', 'г.')


def get_store_name(branch_1c: str):
    conn = psycopg2.connect(host=adb_ip, port=adb_port, database=adb_db_name, user=adb_db_username, password=adb_db_password)
    table_create_query = f'''
            select distinct(store_name) from dwh_data.dim_store where store_name like '%Торговый%'and name_1c_zup = '{branch_1c.strip()}'
            and current_date between datestart and dateend
            '''
    cur = conn.cursor()
    cur.execute(table_create_query)

    df_ = pd.DataFrame(cur.fetchall())

    return df_[df_.columns[0]].iloc[0]


def get_single_report(short_name_: str, name_up: str, name_down: str):

    return ['success', os.path.join(saving_path_1c, f'{short_name_}.xlsx'), '']

    try:

        app = Odines()
        app.run()

        print('started navigating')

        app.navigate("Файл", "Открыть...")

        print('navigated')

        app1 = App('')
        app1.wait_element({"title": "Открытие", "class_name": "#32770", "control_type": "Window",
                           "visible_only": True, "enabled_only": True, "found_index": 0})

        app1.find_element({"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit",
                           "visible_only": True, "enabled_only": True, "found_index": 0}).click()

        app1.find_element({"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit",
                           "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(r'\\172.16.8.87\d\.rpa\.agent\robot-stat-1t\РегламентированныйОтчетФорма1ТКвартальная_на тест.erf', app.keys.ENTER)

        app.parent_switch({"title": "", "class_name": "", "control_type": "Pane", "visible_only": True, "enabled_only": True, "found_index": 20})

        first_input = app.find_element({"title": "", "class_name": "", "control_type": "Edit",
                                        "visible_only": True, "enabled_only": True, "found_index": 0})

        second_input = app.find_element({"title": "", "class_name": "", "control_type": "Edit",
                                         "visible_only": True, "enabled_only": True, "found_index": 1})

        first_input.click()

        sleep(.1)

        keyboard.send_keys("%+r")

        app.parent_switch(app.root)

        # ? ---
        # app.parent_switch({"title": "", "class_name": "", "control_type": "Pane", "visible_only": True, "enabled_only": True, "found_index": 25})

        # arrow_right = app.find_element({"title": "", "class_name": "", "control_type": "Button",
        #                                 "visible_only": True, "enabled_only": True, "found_index": 3}, timeout=1)

        # all_branches = get_all_branches()

        print(name_down, '|||', short_name_)

        print(name_up, name_down, sep=' | ', end='')

        app.find_element({"title": "", "class_name": "", "control_type": "Button",
                          "visible_only": True, "enabled_only": True, "found_index": 2}, timeout=1).click()

        first_input.click()
        first_input.type_keys("^a")
        first_input.type_keys("{BACKSPACE}")
        first_input.type_keys(name_up.strip(), protect_first=True)

        second_input.click()

        if app.wait_element({"title": "1С:Предприятие", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
                             "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=1.5):
            app.find_element({"title": "Нет", "class_name": "", "control_type": "Button", "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            print(f' - BAD', end='')
            logger.info(f'{name_up} | {name_down} - BAD')

        else:
            print(f' - GOOD', end='')
            logger.info(f'{name_up} | {name_down} - GOOD')

        second_input.type_keys("^a")
        second_input.type_keys("{BACKSPACE}")
        second_input.type_keys(name_down.strip(), protect_first=True)

        if app.wait_element({"title": "1С:Предприятие", "class_name": "V8NewLocalFrameBaseWnd", "control_type": "Window",
                             "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=1.5):
            app.find_element({"title": "Нет", "class_name": "", "control_type": "Button", "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            print(f' - BAD",')
            logger.info(f'{name_up} | {name_down} - BAD')

        else:
            print(f' - GOOD",')
            logger.info(f'{name_up} | {name_down} - GOOD')
        print()
        # return ['sucess', '', '']
        app.find_element({"title": "ОК", "class_name": "", "control_type": "Button", "visible_only": True, "enabled_only": True, "found_index": 0}).click()

        app.find_element({"title": "Заполнить", "class_name": "", "control_type": "Button", "visible_only": True, "enabled_only": True, "found_index": 0}).click()

        checker = False

        for _ in range(1000):
            with suppress(Exception):
                app.find_element({"title": "", "class_name": "", "control_type": "DataGrid", "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=10).click()
                checker = True
                break
            sleep(10)

        if checker:

            app.navigate("Файл", "Сохранить как...")

            app.wait_element({"title": "Сохранение", "class_name": "#32770", "control_type": "Window",
                              "visible_only": True, "enabled_only": True, "found_index": 0})

            app.find_element({"title": "Тип файла:", "class_name": "AppControlHost", "control_type": "ComboBox",
                              "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            app.find_element({"title": "Лист Excel2007-... (*.xlsx)", "class_name": "", "control_type": "ListItem",
                              "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            app.find_element({"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit",
                              "visible_only": True, "enabled_only": True, "found_index": 0}).click()

            app.find_element({"title": "Имя файла:", "class_name": "Edit", "control_type": "Edit",
                              "visible_only": True, "enabled_only": True, "found_index": 0}).type_keys(os.path.join(saving_path_1c, short_name_), app.keys.ENTER)

            doc_already_exists = app.wait_element(
                {"title": "Подтвердить сохранение в виде", "class_name": "#32770", "control_type": "Window",
                 "visible_only": True, "enabled_only": True, "found_index": 0}, timeout=2)

            if doc_already_exists:
                app.find_element(
                    {"title": "Да", "class_name": "CCPushButton", "control_type": "Button", "visible_only": True,
                     "enabled_only": True, "found_index": 0}).click()
                time.sleep(0.3)
                if app.wait_element(
                        {"title": "Да", "class_name": "CCPushButton", "control_type": "Button", "visible_only": True,
                         "enabled_only": True, "found_index": 0}, timeout=2):
                    app.find_element(
                        {"title": "Да", "class_name": "CCPushButton", "control_type": "Button", "visible_only": True,
                         "enabled_only": True, "found_index": 0}).click()

            return ['success', os.path.join(saving_path_1c, f'{short_name_}.xlsx'), '']

    except Exception as err:
        return ['error', '', str(err)]


def edit_main_excel_file(filepath__: str, branch_name: str, filepath: str, number: str):

    os.system('taskkill /im excel.exe /f')
    print('OPENING EXCEL', filepath__)
    main_excel = xw.Book(filepath__, corrupt_load=True)

    needed_sheet_name = None

    for sheet_name in main_excel.sheets:
        # print(number, sheet_name.name)
        if number.replace('№', '').replace(' ', '') == sheet_name.name.replace('№', '').replace(' ', ''):
            needed_sheet_name = sheet_name
            break

    print('FOUND SHEET:', branch_name, needed_sheet_name, filepath__, sep=' | ')

    if needed_sheet_name is None:
        main_excel.close()
        return None
    else:
        # needed_sheet_name = needed_sheet_name.name
        main_excel.close()
        #
        # return needed_sheet_name
        return 'H'
    main_sheet = main_excel.sheets[needed_sheet_name]

    quarter = None

    var_4 = '4'
    var_21 = '21'
    var_29 = '29'
    var_31 = '31'

    if main_sheet[f'A34'].value is None:
        var_4 = '3'
        var_21 = '20'
        var_29 = '28'
        var_31 = '30'

    for col_ in 'BDFH':
        if main_sheet[f'{col_}{var_4}'].value is None or main_sheet[f'{col_}{var_21}'].value is None:
            quarter = col_
            break

    if quarter is None:
        main_excel.close()
        return None

    print('col:', quarter)
    print(quarter, var_4, var_21, var_29, var_31)
    print(main_sheet[f'{quarter}{var_4}'].value, main_sheet[f'{quarter}{var_21}'].value, type(main_sheet[f'{quarter}{var_4}'].value), type(main_sheet[f'{quarter}{var_21}'].value), sep=' | ')

    print(f"{main_sheet[f'{quarter}{var_31}'].value} | main_sheet[f'{quarter}{var_31}'].value")

    workers_end_of_period_main = int(main_sheet[f'{quarter}{var_31}'].value)

    excel_app = xw.App(visible=False)
    excel_app.books.open(filepath, corrupt_load=True)

    app = xw.apps.active
    # branch_excel = xlrd.open_workbook(filepath)
    # print(app.range('AJ155').value)

    values = {
        '4': 'AJ92',
        '5': 'AJ96',
        '6': 'AJ98',
        '7': 'AJ102',
        '8': 'AJ104',
        '9': 'AJ110',
        '12': 'AJ116',
        '13': 'AJ118',
        '14': 'AJ120'
    }

    workers_hired: int = int(app.range('AJ132').value) if app.range('AJ132').value is not None else 0
    workers_end_of_period: int = int(app.range('AJ155').value) if app.range('AJ155').value is not None else 0
    workers_fired = int(sum([s for s in app.range('AJ138:AJ153').value if s is not None]))
    # print(workers_fired)
    # print(workers_hired, workers_end_of_period, workers_end_of_period_main, workers_fired)
    if workers_end_of_period != (workers_end_of_period_main + workers_hired - workers_fired):
        # print(int(workers_end_of_period), int((workers_end_of_period_main + workers_hired - workers_fired)))
        workers_hired += int(workers_end_of_period) - int((workers_end_of_period_main + workers_hired - workers_fired))

    # print(workers_hired, workers_end_of_period, workers_end_of_period_main, workers_fired)

    for key_, val in values.items():
        key = int(key_)
        if main_sheet[f'A34'].value is None:
            key -= 1
        if key != 8 or key != 9:
            if app.range(val).value is not None:
                main_sheet[f'{quarter}{key}'].value = round(float(app.range(val).value))
            else:
                main_sheet[f'{quarter}{key}'].value = 0
        else:
            main_sheet[f'{quarter}{key}'].value = app.range(val).value

    main_sheet[f'{quarter}{var_21}'].value = workers_hired
    main_sheet[f'{quarter}{var_29}'].value = workers_fired

    main_excel.app.calculate()
    print(os.path.join(filled_files, filepath__))
    main_excel.save(os.path.join(filled_files, filepath__))
    main_excel.close()

    os.system('taskkill /im excel.exe /f')

    return quarter


def open_1c_zup():

    all_branches = pd.read_excel(mapping_file)
    c = 0

    all_excels_ = []

    print()

    pattern = r'\d+'

    # ! TODO
    # ! branches_to_execute

    all_branches = all_branches.drop_duplicates()

    checkus = False
    aa = []
    for ind in range(len(all_branches)):
        print(f"Started {all_branches['Низ'].iloc[ind]}", end=' ')
        try:
            short_name_ = get_store_name(all_branches['Низ'].iloc[ind])
        except:
            try:
                short_name_ = get_store_name(all_branches['Низ'].iloc[ind].replace('С', 'C'))
            except Exception as errorkin:
                print(f"Branch {all_branches['Низ'].iloc[ind]} is dead: {errorkin}")
                logger.info(f"{datetime.datetime.now()} | Branch {all_branches['Низ'].iloc[ind]} is dead: {errorkin}")
                continue

        if short_name_ == 'Торговый зал СТМ 5АСФ':
            short_name_ = 'Торговый зал АСФ №1'

        print(' | ', short_name_)

        branches_to_execute = get_data_to_execute()

        # if short_name_.replace('Торговый зал ', '') not in list(branches_to_execute['short_name']):
        #     print('skipped', short_name_)
        #     continue

        numbers = re.findall(pattern, all_branches['Низ'].iloc[ind])
        # with suppress(Exception):
        #     if numbers[0] == '27':
        #         print(all_branches['Низ'].iloc[ind])
        if True:
            # # if 'алмат' in all_branches['Низ'].iloc[ind].lower() and int(numbers[0]) <= 39 and 'центр' not in all_branches['Низ'].iloc[ind].lower():
            # print(numbers, '|||', all_branches['Низ'].iloc[ind])
            # if all_branches['Низ'].iloc[ind] == 'Алматинский филиал №1 ТОО "Magnum Cash&Carry"' or all_branches['Низ'].iloc[ind] == 'Алматинский филиал №10 ТОО "Magnum Cash&Carry"' or all_branches['Низ'].iloc[ind] == 'Алматинский филиал №11 ТОО "Magnum Cash&Carry"'\
            #    or all_branches['Низ'].iloc[ind] == 'Алматинский филиал №12 ТОО "Magnum Cash&Carry"' or all_branches['Низ'].iloc[ind] == 'Алматинский филиал №13 ТОО "Magnum Cash&Carry"' or all_branches['Низ'].iloc[ind] == 'Алматинский филиал №14 ТОО "Magnum Cash&Carry"'\
            #    or all_branches['Низ'].iloc[ind] == 'Алматинский филиал №15 ТОО "Magnum Cash&Carry"':
            #     continue
            #
            # # if all_branches['Низ'].iloc[ind] != 'Алматинский филиал №9 ТОО "Magnum Cash&Carry"':
            # #     continue
            # if 'Алмат' in all_branches['Низ'].iloc[ind]:
            #     continue
            # if 'Астан' in all_branches['Низ'].iloc[ind] or 'Султан' in all_branches['Низ'].iloc[ind]:
            #     aa.append(short_name_)
            #     continue

            # * -----
            # checkus = True
            # for file in os.listdir(saving_path_1c):
            #     if short_name_ == file.replace('.xlsx', ''):
            #         # print(f'SKIPPING FILE: {file}')
            #         aa.append(short_name_)
            #         checkus = False
            #         break
            #
            # if checkus:
            #     continue
            # * -----

            # aa.append(short_name_)
            # continue
            # if short_name_ not in ['Торговый зал АФ №14', 'Торговый зал АФ №15', 'Торговый зал АФ №13', 'Торговый зал АФ №36', 'Торговый зал АФ №37', 'Торговый зал АФ №40', 'Торговый зал АФ №41', 'Торговый зал АФ №42', 'Торговый зал АФ №43', 'Торговый зал АФ №44', 'Торговый зал АФ №45', 'Торговый зал АФ №46', 'Торговый зал АФ №47', 'Торговый зал АФ №48', 'Торговый зал АФ №49', 'Торговый зал АФ №50', 'Торговый зал АФ №51', 'Торговый зал АФ №52', 'Торговый зал АФ №53', 'Торговый зал АФ №54', 'Торговый_зал АФ №55 ОПТ', 'Торговый зал АФ №56', 'Торговый зал АФ №57', 'Торговый зал АФ №58', 'Торговый зал АФ №59', 'Торговый зал АФ №60', 'Торговый зал АФ №61', 'Торговый зал АФ №62', 'Торговый зал АФ №63', 'Торговый зал АФ №64', 'Торговый зал АФ №65', 'Торговый зал АФ №66', 'Торговый зал АФ №67', 'Торговый зал АФ №68', 'Торговый зал АФ №69', 'Торговый зал АФ №70', 'Торговый зал АФ №72', 'Торговый зал АФ №73', 'Торговый зал АФ №75', 'Торговый зал АФ №76', 'Торговый зал АФ №77', 'Торговый зал АФ №78', 'Торговый зал АФ №80', 'Торговый зал АФ №81', 'Торговый зал АФ №82', 'Торговый зал АФ №83', 'Торговый зал АФ №84', 'Торговый зал АФ №86', 'Торговый зал СТМ 1АФ']:
            #     continue
            start_time_ = time.time()

            insert_data_in_db(started_time=datetime.datetime.now()   .strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=str(all_branches['Низ'].iloc[ind]), short_name=str(short_name_).replace('Торговый зал ', ''),
                              executor_name=str(ip_address), status_='', status_1c='processing', error_reason='', error_saved_path='', execution_time=0, ecp_path_=os.path.join(ecp_paths, all_branches['Низ'].iloc[ind]))
            try:
                status_, filepath, error_ = get_single_report(short_name_, all_branches['Верх'].iloc[ind], all_branches['Низ'].iloc[ind])

                insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=all_branches['Низ'].iloc[ind], short_name=short_name_.replace('Торговый зал ', ''),
                                  executor_name=ip_address, status_='', status_1c=status_, error_reason=error_, error_saved_path='', execution_time=round(time.time() - start_time_), ecp_path_=os.path.join(ecp_paths, all_branches['Низ'].iloc[ind]))

                # filepath = ind
                all_excels_.append({short_name_: [filepath, short_name_.replace('Торговый зал ', ''), numbers[0]]})

            except Exception as error__:

                insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=all_branches['Низ'].iloc[ind], short_name=short_name_.replace('Торговый зал ', ''),
                                  executor_name=ip_address, status_='', status_1c='failed 1C', error_reason=str(error__), error_saved_path='', execution_time=round(time.time() - start_time_), ecp_path_=os.path.join(ecp_paths, all_branches['Низ'].iloc[ind]))

    print(len(aa))
    print(aa)
    # sleep(10000)

    return all_excels_


def get_data_to_fill(filepath: str, short_name_: str, col_: str):

    os.system('taskkill /im excel.exe /f')

    main_excel = xw.Book(filepath, corrupt_load=True, read_only=True) # * os.path.join(saving_path, '1Т Заполненный файл.xlsx')

    needed_sheet_name = None
    print('col:', col_)
    print('short_name:', short_name_)
    number = short_name_.split('№')[1]

    for sheet_name in main_excel.sheets:
        # if number == sheet_name.name.split()[1]:
        if short_name_.replace('Торговый зал', '').replace('№', '').replace(' ', '') == sheet_name.name.replace('Торговый зал', '').replace('№', '').replace(' ', ''):
            needed_sheet_name = sheet_name
            break

    print()
    print(short_name_.replace('Торговый зал ', '').replace(' №', ''))
    print(needed_sheet_name.name)
    sheet_name = needed_sheet_name.name
    main_sheet = main_excel.sheets[needed_sheet_name]

    first_part_, second_part_ = dict(), dict()

    first_vals = {
        0: 1,
        1: 2,
        2: 3,
        3: 4,
        4: 5,
        5: 6,
        6: 9,
        7: 10,
        8: 11
    }

    second_vals = {
        0: 2,
        1: 10,
        2: 11,
        3: 12
    }

    for ind_, row_ in enumerate(['4', '5', '6', '7', '8', '9', '12', '13', '14']):

        row = int(row_)

        if main_sheet[f'A34'].value is None:
            row -= 1

        if row == 8 or row == 9:

            first_part_.update({first_vals.get(ind_): [main_sheet.range(f'{col_}{row}').value, main_sheet.range(f'{chr(ord(col_) + 1)}{row}').value]})

        else:
            left = round(main_sheet.range(f'{col_}{row}').value, 1) if main_sheet.range(f'{col_}{row}').value is not None else None
            right = round(main_sheet.range(f'{chr(ord(col_) + 1)}{row}').value, 1) if main_sheet.range(f'{chr(ord(col_) + 1)}{row}').value is not None else None

            first_part_.update({first_vals.get(ind_): [left, right]})

    maxx_ = 0

    for i in range(1, 5):
        with suppress(Exception):
            maxx_ = max(maxx_, first_part_.get(i)[0])
    # print('MAXIMUS:', maxx_)
    for i in range(1, 5):
        # print({first_part_.get(i)[0]: [maxx_, first_part_.get(i)[1]]})
        first_part_.update({i: [maxx_, first_part_.get(i)[1]]})
    #
    # if first_part_.get(5)[1] != first_part_.get(6)[1]:
    #     first_part_.update({5: [first_part_.get(5)[0], first_part_.get(6)[1]]})

    for ind_, row_ in enumerate(['21', '29', '30', '31']):

        row = int(row_)

        if main_sheet[f'A34'].value is None:
            row -= 1

        left = round(main_sheet.range(f'{col_}{row}').value, 1) if main_sheet.range(f'{col_}{row}').value is not None else None
        right = round(main_sheet.range(f'{chr(ord(col_) + 1)}{row}').value, 1) if main_sheet.range(f'{chr(ord(col_) + 1)}{row}').value is not None else None

        second_part_.update({second_vals.get(ind_): [left, right]})

    main_excel.close()

    os.system('taskkill /im excel.exe /f')

    return sheet_name, first_part_, second_part_


def dispatcher():

    pass

    # df = pd.read_excel(main_excel_file)


if __name__ == '__main__':

    if True:

        sql_create_table()

        dispatcher()

        # all_excels = open_1c_zup()

        all_excels = [{'Торговый зал АФ №5': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №5.xlsx', 'АФ №5', '5']}, {'Торговый зал АФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №1.xlsx', 'АФ №1', '1']}, {'Торговый зал АФ №10': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №10.xlsx', 'АФ №10', '10']}, {'Торговый зал АФ №11': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №11.xlsx', 'АФ №11', '11']}, {'Торговый зал АФ №12': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №12.xlsx', 'АФ №12', '12']}, {'Торговый зал АФ №14': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №14.xlsx', 'АФ №14', '14']}, {'Торговый зал АФ №15': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №15.xlsx', 'АФ №15', '15']}, {'Торговый зал АФ №16': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №16.xlsx', 'АФ №16', '16']}, {'Торговый зал АФ №17': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №17.xlsx', 'АФ №17', '17']}, {'Торговый зал АФ №18': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №18.xlsx', 'АФ №18', '18']}, {'Торговый зал АФ №19': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №19.xlsx', 'АФ №19', '19']}, {'Торговый зал АФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №2.xlsx', 'АФ №2', '2']}, {'Торговый зал АФ №20': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №20.xlsx', 'АФ №20', '20']}, {'Торговый зал АФ №21': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №21.xlsx', 'АФ №21', '21']}, {'Торговый зал АФ №22': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №22.xlsx', 'АФ №22', '22']}, {'Торговый зал АФ №23': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №23.xlsx', 'АФ №23', '23']}, {'Торговый зал АФ №24': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №24.xlsx', 'АФ №24', '24']}, {'Торговый зал АФ №25': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №25.xlsx', 'АФ №25', '25']}, {'Торговый зал АФ №26': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №26.xlsx', 'АФ №26', '26']}, {'Торговый зал АФ №28': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №28.xlsx', 'АФ №28', '28']}, {'Торговый зал АФ №29': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №29.xlsx', 'АФ №29', '29']}, {'Торговый зал АФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №3.xlsx', 'АФ №3', '3']}, {'Торговый зал АФ №30': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №30.xlsx', 'АФ №30', '30']}, {'Торговый зал АФ №31': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №31.xlsx', 'АФ №31', '31']}, {'Торговый зал АФ №32': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №32.xlsx', 'АФ №32', '32']}, {'Торговый зал АФ №33': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №33.xlsx', 'АФ №33', '33']}, {'Торговый зал АФ №34': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №34.xlsx', 'АФ №34', '34']}, {'Торговый зал АФ №35': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №35.xlsx', 'АФ №35', '35']}, {'Торговый зал АФ №6': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №6.xlsx', 'АФ №6', '6']}, {'Торговый зал АФ №7': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №7.xlsx', 'АФ №7', '7']}, {'Торговый зал АФ №8': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №8.xlsx', 'АФ №8', '8']}, {'Торговый зал АФ №9': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №9.xlsx', 'АФ №9', '9']}, {'Торговый зал КФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КФ №2.xlsx', 'КФ №2', '2']}, {'Торговый зал ЕКФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ЕКФ №1.xlsx', 'ЕКФ №1', '1']}, {'Торговый зал КПФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КПФ №1.xlsx', 'КПФ №1', '1']}, {'Торговый зал ФКС №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ФКС №1.xlsx', 'ФКС №1', '1']}, {'Торговый зал КЗФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КЗФ №1.xlsx', 'КЗФ №1', '1']}, {'Торговый зал ТФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТФ №1.xlsx', 'ТФ №1', '1']}, {'Торговый зал ППФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №1.xlsx', 'ППФ №1', '1']}, {'Торговый зал ТЗФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТЗФ №1.xlsx', 'ТЗФ №1', '1']}, {'Торговый зал УКФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал УКФ №1.xlsx', 'УКФ №1', '1']}, {'Торговый зал ППФ №10': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №10.xlsx', 'ППФ №10', '10']}, {'Торговый зал ШФ №10': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №10.xlsx', 'ШФ №10', '10']}, {'Торговый зал ППФ №11': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №11.xlsx', 'ППФ №11', '11']}, {'Торговый зал ППФ №12': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №12.xlsx', 'ППФ №12', '12']}, {'Торговый зал ШФ №12': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №12.xlsx', 'ШФ №12', '12']}, {'Торговый зал ППФ №13': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №13.xlsx', 'ППФ №13', '13']}, {'Торговый зал ШФ №13': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №13.xlsx', 'ШФ №13', '13']}, {'Торговый зал АФ №13': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №13.xlsx', 'АФ №13', '13']}, {'Торговый зал ППФ №14': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №14.xlsx', 'ППФ №14', '14']}, {'Торговый зал ШФ №14': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №14.xlsx', 'ШФ №14', '14']}, {'Торговый зал ППФ №15': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №15.xlsx', 'ППФ №15', '15']}, {'Торговый зал ШФ №15': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №15.xlsx', 'ШФ №15', '15']}, {'Торговый зал АСФ №16': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №16.xlsx', 'АСФ №16', '16']}, {'Торговый зал ППФ №16': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №16.xlsx', 'ППФ №16', '16']}, {'Торговый зал ППФ №17': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №17.xlsx', 'ППФ №17', '17']}, {'Торговый зал ШФ №17': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №17.xlsx', 'ШФ №17', '17']}, {'Торговый зал АСФ №17': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №17.xlsx', 'АСФ №17', '17']}, {'Торговый зал АСФ №18': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №18.xlsx', 'АСФ №18', '18']}, {'Торговый зал ППФ №18': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №18.xlsx', 'ППФ №18', '18']}, {'Торговый зал ШФ №18': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №18.xlsx', 'ШФ №18', '18']}, {'Торговый зал АСФ №19': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №19.xlsx', 'АСФ №19', '19']}, {'Торговый зал ППФ №19': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №19.xlsx', 'ППФ №19', '19']}, {'Торговый зал ШФ №19': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №19.xlsx', 'ШФ №19', '19']}, {'Торговый зал ФКС №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ФКС №2.xlsx', 'ФКС №2', '2']}, {'Торговый зал КЗФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КЗФ №2.xlsx', 'КЗФ №2', '2']}, {'Торговый зал ППФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №2.xlsx', 'ППФ №2', '2']}, {'Торговый зал ТКФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТКФ №2.xlsx', 'ТКФ №2', '2']}, {'Торговый зал ТЗФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТЗФ №2.xlsx', 'ТЗФ №2', '2']}, {'Торговый зал ТФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТФ №2.xlsx', 'ТФ №2', '2']}, {'Торговый зал АСФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №2.xlsx', 'АСФ №2', '2']}, {'Торговый зал УКФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал УКФ №2.xlsx', 'УКФ №2', '2']}, {'Торговый зал ШФ №2': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №2.xlsx', 'ШФ №2', '2']}, {'Торговый зал АСФ №20': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №20.xlsx', 'АСФ №20', '20']}, {'Торговый зал ППФ №20': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №20.xlsx', 'ППФ №20', '20']}, {'Торговый зал ШФ №20': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №20.xlsx', 'ШФ №20', '20']}, {'Торговый зал АСФ №21': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №21.xlsx', 'АСФ №21', '21']}, {'Торговый зал ППФ №21': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №21.xlsx', 'ППФ №21', '21']}, {'Торговый зал ШФ №21': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №21.xlsx', 'ШФ №21', '21']}, {'Торговый зал АСФ №22': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №22.xlsx', 'АСФ №22', '22']}, {'Торговый зал ППФ №22': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №22.xlsx', 'ППФ №22', '22']}, {'Торговый зал ШФ №22': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №22.xlsx', 'ШФ №22', '22']}, {'Торговый зал АСФ №23': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №23.xlsx', 'АСФ №23', '23']}, {'Торговый зал ШФ №23': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №23.xlsx', 'ШФ №23', '23']}, {'Торговый зал ШФ №24': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №24.xlsx', 'ШФ №24', '24']}, {'Торговый зал АСФ №24': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №24.xlsx', 'АСФ №24', '24']}, {'Торговый зал АСФ №25': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №25.xlsx', 'АСФ №25', '25']}, {'Торговый зал ШФ №25': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №25.xlsx', 'ШФ №25', '25']}, {'Торговый зал АСФ №26': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №26.xlsx', 'АСФ №26', '26']}, {'Торговый зал ШФ №26': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №26.xlsx', 'ШФ №26', '26']}, {'Торговый зал АСФ №27': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №27.xlsx', 'АСФ №27', '27']}, {'Торговый зал ШФ №27': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №27.xlsx', 'ШФ №27', '27']}, {'Торговый зал ШФ №28': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №28.xlsx', 'ШФ №28', '28']}, {'Торговый зал АСФ №28': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №28.xlsx', 'АСФ №28', '28']}, {'Торговый зал АСФ №29': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №29.xlsx', 'АСФ №29', '29']}, {'Торговый зал ШФ №29': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №29.xlsx', 'ШФ №29', '29']}, {'Торговый зал ТФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТФ №3.xlsx', 'ТФ №3', '3']}, {'Торговый зал АСФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №3.xlsx', 'АСФ №3', '3']}, {'Торговый зал КФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КФ №3.xlsx', 'КФ №3', '3']}, {'Торговый зал ППФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №3.xlsx', 'ППФ №3', '3']}, {'Торговый зал ТЗФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТЗФ №3.xlsx', 'ТЗФ №3', '3']}, {'Торговый зал УКФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал УКФ №3.xlsx', 'УКФ №3', '3']}, {'Торговый зал ШФ №3': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №3.xlsx', 'ШФ №3', '3']}, {'Торговый зал АСФ №30': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №30.xlsx', 'АСФ №30', '30']}, {'Торговый зал ШФ №30': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №30.xlsx', 'ШФ №30', '30']}, {'Торговый зал АСФ №31': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №31.xlsx', 'АСФ №31', '31']}, {'Торговый зал АСФ №32': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №32.xlsx', 'АСФ №32', '32']}, {'Торговый зал ШФ №32': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №32.xlsx', 'ШФ №32', '32']}, {'Торговый зал АСФ №33': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №33.xlsx', 'АСФ №33', '33']}, {'Торговый зал ШФ №33': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №33.xlsx', 'ШФ №33', '33']}, {'Торговый зал АСФ №34': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №34.xlsx', 'АСФ №34', '34']}, {'Торговый зал ШФ №34': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №34.xlsx', 'ШФ №34', '34']}, {'Торговый зал АСФ №35': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №35.xlsx', 'АСФ №35', '35']}, {'Торговый зал_ОПТ ШФ №35': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал_ОПТ ШФ №35.xlsx', 'Торговый зал_ОПТ ШФ №35', '35']}, {'Торговый зал АСФ №36': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №36.xlsx', 'АСФ №36', '36']}, {'Торговый зал АФ №36': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №36.xlsx', 'АФ №36', '36']}, {'Торговый зал АСФ №37': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №37.xlsx', 'АСФ №37', '37']}, {'Торговый зал АФ №37': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №37.xlsx', 'АФ №37', '37']}, {'Торговый зал АСФ №38': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №38.xlsx', 'АСФ №38', '38']}, {'Торговый зал АФ №38': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №38.xlsx', 'АФ №38', '38']}, {'Торговый зал АСФ №39': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №39.xlsx', 'АСФ №39', '39']}, {'Торговый зал АФ №39': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №39.xlsx', 'АФ №39', '39']}, {'Торговый зал АСФ №4': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №4.xlsx', 'АСФ №4', '4']}, {'Торговый зал АФ №4': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №4.xlsx', 'АФ №4', '4']}, {'Торговый зал КФ №4': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КФ №4.xlsx', 'КФ №4', '4']}, {'Торговый зал ППФ №4': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №4.xlsx', 'ППФ №4', '4']}, {'Торговый зал ТФ №4': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ТФ №4.xlsx', 'ТФ №4', '4']}, {'Торговый зал ШФ №4': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №4.xlsx', 'ШФ №4', '4']}, {'Торговый зал АСФ №40': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №40.xlsx', 'АСФ №40', '40']}, {'Торговый зал АФ №40': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №40.xlsx', 'АФ №40', '40']}, {'Торговый зал АСФ №41': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №41.xlsx', 'АСФ №41', '41']}, {'Торговый зал АФ №41': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №41.xlsx', 'АФ №41', '41']}, {'Торговый зал АСФ №42': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №42.xlsx', 'АСФ №42', '42']}, {'Торговый зал АФ №42': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №42.xlsx', 'АФ №42', '42']}, {'Торговый зал АФ №43': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №43.xlsx', 'АФ №43', '43']}, {'Торговый зал АСФ №44': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №44.xlsx', 'АСФ №44', '44']}, {'Торговый зал АФ №44': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №44.xlsx', 'АФ №44', '44']}, {'Торговый зал АСФ №45': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №45.xlsx', 'АСФ №45', '45']}, {'Торговый зал АФ №45': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №45.xlsx', 'АФ №45', '45']}, {'Торговый зал АСФ №46': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №46.xlsx', 'АСФ №46', '46']}, {'Торговый зал АФ №46': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №46.xlsx', 'АФ №46', '46']}, {'Торговый зал АСФ №47': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №47.xlsx', 'АСФ №47', '47']}, {'Торговый зал АФ №47': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №47.xlsx', 'АФ №47', '47']}, {'Торговый зал АСФ №48': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №48.xlsx', 'АСФ №48', '48']}, {'Торговый зал АФ №48': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №48.xlsx', 'АФ №48', '48']}, {'Торговый зал АФ №49': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №49.xlsx', 'АФ №49', '49']}, {'Торговый зал АСФ №5': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №5.xlsx', 'АСФ №5', '5']}, {'Торговый зал КФ №5': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КФ №5.xlsx', 'КФ №5', '5']}, {'Торговый зал ППФ №5': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №5.xlsx', 'ППФ №5', '5']}, {'Торговый зал ШФ №5': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №5.xlsx', 'ШФ №5', '5']}, {'Торговый зал АСФ №50': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №50.xlsx', 'АСФ №50', '50']}, {'Торговый зал АФ №50': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №50.xlsx', 'АФ №50', '50']}, {'Торговый зал АСФ №51': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №51.xlsx', 'АСФ №51', '51']}, {'Торговый зал АФ №51': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №51.xlsx', 'АФ №51', '51']}, {'Торговый зал АСФ №52': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №52.xlsx', 'АСФ №52', '52']}, {'Торговый зал АФ №52': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №52.xlsx', 'АФ №52', '52']}, {'Торговый зал АСФ №53': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №53.xlsx', 'АСФ №53', '53']}, {'Торговый зал АФ №53': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №53.xlsx', 'АФ №53', '53']}, {'Торговый зал АСФ №54': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №54.xlsx', 'АСФ №54', '54']}, {'Торговый зал АФ №54': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №54.xlsx', 'АФ №54', '54']}, {'Торговый зал АСФ №55': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №55.xlsx', 'АСФ №55', '55']}, {'Торговый_зал АФ №55 ОПТ': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый_зал АФ №55 ОПТ.xlsx', 'Торговый_зал АФ №55 ОПТ', '55']}, {'Торговый зал АСФ №56': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №56.xlsx', 'АСФ №56', '56']}, {'Торговый зал АФ №56': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №56.xlsx', 'АФ №56', '56']}, {'Торговый зал АСФ №57': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №57.xlsx', 'АСФ №57', '57']}, {'Торговый зал АФ №57': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №57.xlsx', 'АФ №57', '57']}, {'Торговый зал АСФ №58': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №58.xlsx', 'АСФ №58', '58']}, {'Торговый зал АФ №58': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №58.xlsx', 'АФ №58', '58']}, {'Торговый зал АСФ №59': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №59.xlsx', 'АСФ №59', '59']}, {'Торговый зал АФ №59': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №59.xlsx', 'АФ №59', '59']}, {'Торговый зал АСФ №6': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №6.xlsx', 'АСФ №6', '6']}, {'Торговый зал КФ №6': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КФ №6.xlsx', 'КФ №6', '6']}, {'Торговый зал ППФ №6': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №6.xlsx', 'ППФ №6', '6']}, {'Торговый зал ШФ №6': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №6.xlsx', 'ШФ №6', '6']}, {'Торговый зал АСФ №60': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №60.xlsx', 'АСФ №60', '60']}, {'Торговый зал АФ №60': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №60.xlsx', 'АФ №60', '60']}, {'Торговый зал АСФ №61': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №61.xlsx', 'АСФ №61', '61']}, {'Торговый зал АФ №61': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №61.xlsx', 'АФ №61', '61']}, {'Торговый зал АСФ №62': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №62.xlsx', 'АСФ №62', '62']}, {'Торговый зал АФ №62': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №62.xlsx', 'АФ №62', '62']}, {'Торговый зал АСФ №63': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №63.xlsx', 'АСФ №63', '63']}, {'Торговый зал АФ №63': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №63.xlsx', 'АФ №63', '63']}, {'Торговый зал АСФ №64': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №64.xlsx', 'АСФ №64', '64']}, {'Торговый зал АФ №64': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №64.xlsx', 'АФ №64', '64']}, {'Торговый зал АСФ №65': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №65.xlsx', 'АСФ №65', '65']}, {'Торговый зал АФ №65': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №65.xlsx', 'АФ №65', '65']}, {'Торговый зал АСФ №66': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №66.xlsx', 'АСФ №66', '66']}, {'Торговый зал АФ №66': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №66.xlsx', 'АФ №66', '66']}, {'Торговый зал АСФ №67': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №67.xlsx', 'АСФ №67', '67']}, {'Торговый зал АФ №67': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №67.xlsx', 'АФ №67', '67']}, {'Торговый зал АСФ №68': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №68.xlsx', 'АСФ №68', '68']}, {'Торговый зал АФ №68': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №68.xlsx', 'АФ №68', '68']}, {'Торговый зал АСФ №69': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №69.xlsx', 'АСФ №69', '69']}, {'Торговый зал АФ №69': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №69.xlsx', 'АФ №69', '69']}, {'Торговый зал КФ №7': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал КФ №7.xlsx', 'КФ №7', '7']}, {'Торговый зал ППФ №7': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №7.xlsx', 'ППФ №7', '7']}, {'Торговый зал ШФ №7': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №7.xlsx', 'ШФ №7', '7']}, {'Торговый зал АФ №70': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №70.xlsx', 'АФ №70', '70']}, {'Торговый зал АСФ №71': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №71.xlsx', 'АСФ №71', '71']}, {'Торговый зал АСФ №72': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №72.xlsx', 'АСФ №72', '72']}, {'Торговый зал АФ №72': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №72.xlsx', 'АФ №72', '72']}, {'Торговый зал АСФ №73': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №73.xlsx', 'АСФ №73', '73']}, {'Торговый зал АФ №73': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №73.xlsx', 'АФ №73', '73']}, {'Торговый зал АСФ №74': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №74.xlsx', 'АСФ №74', '74']}, {'Торговый зал АСФ №75': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №75.xlsx', 'АСФ №75', '75']}, {'Торговый зал АФ №75': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №75.xlsx', 'АФ №75', '75']}, {'Торговый зал АСФ №76': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №76.xlsx', 'АСФ №76', '76']}, {'Торговый зал АФ №76': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №76.xlsx', 'АФ №76', '76']}, {'Торговый зал АСФ №77': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №77.xlsx', 'АСФ №77', '77']}, {'Торговый зал АФ №77': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №77.xlsx', 'АФ №77', '77']}, {'Торговый зал АФ №78': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №78.xlsx', 'АФ №78', '78']}, {'Торговый зал АСФ №79': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №79.xlsx', 'АСФ №79', '79']}, {'Торговый зал АСФ №8': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №8.xlsx', 'АСФ №8', '8']}, {'Торговый зал ППФ №8': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №8.xlsx', 'ППФ №8', '8']}, {'Торговый зал ШФ №8': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №8.xlsx', 'ШФ №8', '8']}, {'Торговый зал АСФ №80': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №80.xlsx', 'АСФ №80', '80']}, {'Торговый зал АФ №80': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №80.xlsx', 'АФ №80', '80']}, {'Торговый зал АСФ №81': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №81.xlsx', 'АСФ №81', '81']}, {'Торговый зал АФ №81': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №81.xlsx', 'АФ №81', '81']}, {'Торговый зал АСФ №82': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №82.xlsx', 'АСФ №82', '82']}, {'Торговый зал АФ №82': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №82.xlsx', 'АФ №82', '82']}, {'Торговый зал АСФ №83': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №83.xlsx', 'АСФ №83', '83']}, {'Торговый зал АФ №83': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №83.xlsx', 'АФ №83', '83']}, {'Торговый зал АФ №84': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №84.xlsx', 'АФ №84', '84']}, {'Торговый зал АФ №86': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №86.xlsx', 'АФ №86', '86']}, {'Торговый зал ШФ №9': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ШФ №9.xlsx', 'ШФ №9', '9']}, {'Торговый зал ППФ №9': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал ППФ №9.xlsx', 'ППФ №9', '9']}, {'Торговый зал АСФ №1': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №1.xlsx', 'АСФ №1', '1']}, {'Торговый зал СТМ 1АФ': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал СТМ 1АФ.xlsx', 'СТМ 1АФ', '3']}, {'Торговый зал АФ №5': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АФ №5.xlsx', 'АФ №5', '5']}, {'Торговый зал АСФ №10': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №10.xlsx', 'АСФ №10', '10']}, {'Торговый зал АСФ №15': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №15.xlsx', 'АСФ №15', '15']}, {'Торговый зал АСФ №11': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №11.xlsx', 'АСФ №11', '11']}, {'Торговый зал АСФ №12': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №12.xlsx', 'АСФ №12', '12']}, {'Торговый зал АСФ №13': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №13.xlsx', 'АСФ №13', '13']}, {'Торговый зал АСФ №14': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №14.xlsx', 'АСФ №14', '14']}, {'Торговый зал АСФ №7': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №7.xlsx', 'АСФ №7', '7']}, {'Торговый зал АСФ №9': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №9.xlsx', 'АСФ №9', '9']}, {'Торговый зал АСФ №70': ['\\\\172.16.8.87\\d\\.rpa\\.agent\\robot-stat-1t\\Output\\Выгрузка 1Т из 1С\\Торговый зал АСФ №70.xlsx', 'АСФ №70', '70']}]

        print(all_excels)
        # logger.info(all_excels)
        logger.warning(all_excels)
        # for excel in all_excels:
        #     for key, val in excel.items():
        #         print(key, val)
            # print(excel)
            # print('---')
        print('-----------------')
        for filepath_ in os.listdir(main_excel_files):

            pass

            # shutil.copy(os.path.join(main_excel_files, filepath_), os.path.join(filled_files, filepath_))

            # print(all_excels)
            # logger.info(all_excels)
            # logger.warning(all_excels)

            # print("FINISHED!!!!")

            # print(all_excels)
            # print('----')
            # for excel in all_excels:
            #     print(excel)

        # if ip_address == '10.70.2.9':
        #     all_excels = all_excels[::2]
        # if ip_address == '10.70.2.2':
        #     all_excels = all_excels[1::2]
        # if ip_address == '10.70.2.11':
        all_excels = all_excels[::]
        # else:
        #     all_excels = all_excels[:len(all_excels) // 2]

        for excel in all_excels:
            for branch, vals in excel.items():
                # if branch == 'Торговый зал АФ №9' or branch == 'Торговый зал АСФ №17' or branch == 'Торговый зал АСФ №20' \
                #    or branch == 'Торговый зал АСФ №11' or branch == 'Торговый зал АСФ №1':
                #     continue
                if branch in ['Торговый зал АФ №33', 'Торговый зал АФ №35', 'Торговый зал АФ №7', 'Торговый зал КФ №2', 'Торговый зал КПФ №1', 'Торговый зал КЗФ №1', 'Торговый зал ППФ №1', 'Торговый зал АФ №13', 'Торговый зал АСФ №17', 'Торговый зал АСФ №19', 'Торговый зал АСФ №20', 'Торговый зал АСФ №22', 'Торговый зал АСФ №24', 'Торговый зал АСФ №26', 'Торговый зал АСФ №28', 'Торговый зал АСФ №3', 'Торговый зал АСФ №31', 'Торговый зал АСФ №33', 'Торговый зал АСФ №35', 'Торговый зал АФ №36', 'Торговый зал АФ №37', 'Торговый зал АФ №38', 'Торговый зал АФ №39', 'Торговый зал ППФ №4', 'Торговый зал ШФ №4', 'Торговый зал АФ №40', 'Торговый зал АФ №41', 'Торговый зал АФ №43', 'Торговый зал АСФ №47', 'Торговый зал ППФ №5', 'Торговый зал АФ №52', 'Торговый_зал АФ №55 ОПТ', 'Торговый зал АФ №56', 'Торговый зал АФ №57', 'Торговый зал АФ №58', 'Торговый зал АФ №59', 'Торговый зал КФ №6', 'Торговый зал ШФ №6', 'Торговый зал АФ №60', 'Торговый зал АФ №61', 'Торговый зал АФ №62', 'Торговый зал АФ №63', 'Торговый зал АФ №64', 'Торговый зал АФ №65', 'Торговый зал АФ №66', 'Торговый зал АФ №67', 'Торговый зал АФ №68', 'Торговый зал КФ №7', 'Торговый зал ШФ №7', 'Торговый зал АСФ №72', 'Торговый зал АСФ №74', 'Торговый зал АСФ №76', 'Торговый зал АСФ №79', 'Торговый зал ППФ №8', 'Торговый зал АСФ №80', 'Торговый зал АСФ №82', 'Торговый зал ШФ №9', 'Торговый зал СТМ 5АСФ']:
                    continue
                if branch in ['Торговый зал УКФ №1', 'Торговый зал ППФ №10', 'Торговый зал ШФ №10', 'Торговый зал ППФ №11', 'Торговый зал ППФ №12', 'Торговый зал ШФ №12', 'Торговый зал ППФ №13', 'Торговый зал ШФ №13', 'Торговый зал ППФ №14', 'Торговый зал ШФ №14', 'Торговый зал ППФ №15', 'Торговый зал ШФ №15', 'Торговый зал ППФ №16', 'Торговый зал ППФ №17', 'Торговый зал ШФ №17', 'Торговый зал АСФ №17', 'Торговый зал ППФ №18', 'Торговый зал ШФ №18', 'Торговый зал ППФ №19', 'Торговый зал ШФ №19', 'Торговый зал ФКС №2', 'Торговый зал КЗФ №2', 'Торговый зал ППФ №2', 'Торговый зал ТКФ №2', 'Торговый зал ТЗФ №2', 'Торговый зал ТФ №2', 'Торговый зал УКФ №2', 'Торговый зал ШФ №2', 'Торговый зал АСФ №20', 'Торговый зал ППФ №20', 'Торговый зал ШФ №20', 'Торговый зал ППФ №21', 'Торговый зал ШФ №21', 'Торговый зал ППФ №22', 'Торговый зал ШФ №22', 'Торговый зал ШФ №23', 'Торговый зал ШФ №24', 'Торговый зал ШФ №25', 'Торговый зал ШФ №26', 'Торговый зал ШФ №27', 'Торговый зал ШФ №28', 'Торговый зал ШФ №29', 'Торговый зал ТФ №3', 'Торговый зал КФ №3', 'Торговый зал ППФ №3', 'Торговый зал ТЗФ №3', 'Торговый зал УКФ №3', 'Торговый зал ШФ №3', 'Торговый зал ШФ №30', 'Торговый зал ШФ №32', 'Торговый зал ШФ №33', 'Торговый зал ШФ №34', 'Торговый зал_ОПТ ШФ №35', 'Торговый зал АФ №4', 'Торговый зал КФ №4']:
                    continue
                if branch in ['Торговый зал ФКС №1']:
                    continue

                # # * Delete Then
                if branch not in ['Торговый зал АСФ №55', 'Торговый зал АСФ №68', 'Торговый зал АСФ №26', 'Торговый зал АСФ №65'
                                  , 'Торговый зал АСФ №67', 'Торговый зал АСФ №66', 'Торговый зал АСФ №11', 'Торговый зал АСФ №64', 'Торговый зал АСФ №42'
                                  , 'Торговый зал АСФ №52', 'Торговый зал АСФ №73', 'Торговый зал АСФ №36', 'Торговый зал АСФ №37', 'Торговый зал АСФ №48', 'Торговый зал АСФ №61', 'Торговый зал АСФ №69'
                                  , 'Торговый зал АСФ №56', 'Торговый зал АСФ №57', 'Торговый зал АСФ №62', 'Торговый зал АСФ №60',
                                  'Торговый зал АФ №73', 'Торговый зал АФ 46', 'Торговый зал АФ №75', 'Торговый зал АФ №80', 'Торговый зал АФ №78', 'Торговый зал ТФ №1', 'Торговый зал ППФ №9']:
                    continue
                if branch != 'Торговый зал АСФ №73':
                    continue
                # if 'АСФ' not in branch:
                #     continue
                checkus = False
                print('KKK', saving_path)
                for sent_report in os.listdir(reports_saving_path):
                    if branch == sent_report.replace('.jpg', ''):
                        checkus = True
                        break

                if checkus:
                    print(f'SKIPPED BRANCHOS: {branch}')
                    # continue

                logger.warning(f'STARTED BRANCHOS: {branch}')
                # continue
                col = None
                filepath_for_report = None

                for filepath_ in os.listdir(main_excel_files):

                    if True:
                        print(os.path.join(filled_files, filepath_), branch, vals[0], vals[1])
                        col = edit_main_excel_file(os.path.join(filled_files, filepath_), branch, vals[0], vals[1])

                        if col is None:
                            continue
                        else:
                            filepath_for_report = filepath_
                            break

                    # except Exception as err:
                    #     logger.info(f"ERRORUS OCC URED: {err} | {branch}")
                    #     logger.warning(f"ERRORUS OCCURED: {err} | {branch}")
                    #     if 'such file' in str(err):
                    #         logger.info(f'SKIPPING THAT BRANCH {branch}')

                print(f"FOUND COL!!! {branch} ||| {col}")
                print('-------------------')
                col = 'H'
                if '~' not in vals and col is not None and filepath_for_report is not None:
                    pass
                    short_name, first_part, second_part = get_data_to_fill(os.path.join(filled_files, filepath_for_report), branch, col)

                    print('------KEKJON------')
                    print(short_name, first_part, second_part, sep='\n')
                    # sleep(1000)
                    start_time = time.time()
                    insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=branch, short_name=short_name,
                                      executor_name=ip_address, status_='processing', status_1c='success', error_reason='', error_saved_path='', execution_time=0, ecp_path_=os.path.join(ecp_paths, branch))
                    try:
                        status, error_saved_path, error = start_single_branch(branch, first_part, second_part)

                        insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=branch, short_name=short_name,
                                          executor_name=ip_address, status_=status, status_1c='success', error_reason=error, error_saved_path=error_saved_path, execution_time=round(time.time() - start_time), ecp_path_=os.path.join(ecp_paths, branch))

                    except Exception as error:

                        insert_data_in_db(started_time=datetime.datetime.now().strftime("%d.%m.%Y %H:%M:%S.%f"), store_name=branch, short_name=short_name,
                                          executor_name=ip_address, status_='failed with error', status_1c='success', error_reason=str(error), error_saved_path='', execution_time=round(time.time() - start_time), ecp_path_=os.path.join(ecp_paths, branch))

    # except Exception as errorik:
    #     print(f'ERROR: {errorik}')
    #     logger.info(f'ERROR: {errorik}')
    #     logger.warning(f'ERROR: {errorik}')


